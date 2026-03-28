from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, Response
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, random, json, io, re, string, time as _time
from datetime import date as _date

from database import db, User, Faculty, Subject, Practical, Question, Result, AttendanceSession, AttendanceRecord

app = Flask(__name__)
app.secret_key = "super-secret-localhost-key-2026"
app.config['TEMPLATES_AUTO_RELOAD'] = True

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

# MySQL / Deployment Database
MYSQL_URI = os.environ.get("DATABASE_URL", "mysql+pymysql://root:root123@localhost/spas_db?charset=utf8mb4")
app.config["SQLALCHEMY_DATABASE_URI"] = MYSQL_URI
db.init_app(app)

with app.app_context():
    db.create_all()

EXAM_DURATION_SECONDS = 15 * 60


# ─── Helper functions ─────────────────────────────────────
def is_student():
    return "roll_no" in session

def is_faculty():
    return "faculty_id" in session

def is_logged_in():
    return is_student() or is_faculty()

def extract_practical_number(name):
    """Extract the practical number from names like 'Practical No: 1 Write a program...'"""
    match = re.search(r'(?i)practical\s*(?:no\.?\s*)?:?\s*(\d+)', name)
    if match:
        return int(match.group(1))
    match = re.search(r'^\s*(\d+)', name)
    if match:
        return int(match.group(1))
    return 9999

def gen_code(n=6):
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=n))

def get_all_practicals_for_subject(subject_name: str) -> list[str]:
    """Get practical names for a subject"""
    subj = Subject.query.filter_by(name=subject_name).first()
    if subj:
        return [p.name for p in Practical.query.filter_by(subject_id=subj.id).order_by(Practical.sort_order).all()]
    return []

def get_all_practicals_flat() -> list[str]:
    """Get all practical names across all subjects"""
    return [p.name for p in Practical.query.order_by(Practical.sort_order).all()]

def get_student_results(roll_no: str) -> list[dict]:
    """Get all results for a student — latest per practical only"""
    results = Result.query.filter_by(roll_no=roll_no).order_by(Result.timestamp.desc()).all()
    # Dedup: keep latest per practical
    seen = set()
    deduped = []
    for r in results:
        if r.practical_name not in seen:
            seen.add(r.practical_name)
            deduped.append(r.to_dict())
    return deduped


# ═══════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        login_type = request.form.get("login_type")
        action = request.form.get("action")

        if login_type == "student":
            if action == "login":
                roll = request.form.get("roll_no", "").strip()
                password = request.form.get("password", "").strip()
                user = User.query.filter_by(roll_no=roll).first()

                if user and user.check_password(password):
                    session["roll_no"] = roll
                    session["full_name"] = user.full_name
                    session["branch"] = user.branch
                    session["year"] = user.year
                    session["batch"] = user.batch
                    session["email"] = user.email
                    session["user_type"] = "student"
                    flash("Login successful.", "success")
                    return redirect(url_for("dashboard"))
                else:
                    flash("Invalid roll number or password.", "error")

            elif action == "register":
                roll = request.form.get("roll_no", "").strip()
                password = request.form.get("password", "").strip()
                full_name = request.form.get("full_name", "").strip()
                branch = request.form.get("branch")
                year = request.form.get("year")
                batch = request.form.get("batch", "1")
                email = request.form.get("email", "").strip()

                if not roll or not password or len(password) < 6:
                    flash("Roll no + password required, password must be at least 6 characters.", "error")
                elif User.query.filter_by(roll_no=roll).first():
                    flash("Roll number already exists.", "error")
                else:
                    u = User(roll_no=roll, full_name=full_name, branch=branch,
                             year=year, batch=batch, email=email)
                    u.set_password(password)
                    db.session.add(u)
                    db.session.commit()
                    flash("Profile created successfully. Now you can login.", "success")

        elif login_type == "faculty":
            if action == "login":
                faculty_id = request.form.get("faculty_id", "").strip()
                password = request.form.get("password", "").strip()
                fac = Faculty.query.filter_by(faculty_id=faculty_id).first()

                if fac and fac.check_password(password):
                    session["faculty_id"] = faculty_id
                    session["full_name"] = fac.full_name
                    session["department"] = fac.department
                    session["email"] = fac.email
                    session["user_type"] = "faculty"
                    flash("Login successful.", "success")
                    return redirect(url_for("faculty_dashboard"))
                else:
                    flash("Invalid faculty ID or password.", "error")

            elif action == "register":
                faculty_id = request.form.get("faculty_id", "").strip()
                password = request.form.get("password", "").strip()
                full_name = request.form.get("full_name", "").strip()
                department = request.form.get("department")
                subjects_list = request.form.getlist("subjects")
                subjects_str = ",".join(subjects_list)
                email = request.form.get("email", "").strip()

                if not faculty_id or not password or len(password) < 6:
                    flash("Faculty ID + password required, password must be at least 6 characters.", "error")
                elif Faculty.query.filter_by(faculty_id=faculty_id).first():
                    flash("Faculty ID already exists.", "error")
                else:
                    f = Faculty(faculty_id=faculty_id, full_name=full_name,
                                department=department, email=email, subjects=subjects_str)
                    f.set_password(password)
                    db.session.add(f)
                    db.session.commit()
                    flash("Faculty profile created successfully. Now you can login.", "success")

    return render_template("login.html", subjects=Subject.query.all())


@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    user_type = request.form.get("user_type")
    email = request.form.get("email", "").strip()

    if not email or not user_type:
        return jsonify({"success": False, "message": "Email and user type required."})

    if user_type == "student":
        user = User.query.filter_by(email=email).first()
        if user:
            pwd = user.plain_password
            if not pwd:
                pwd = "spas_" + gen_code(4)
                user.set_password(pwd)
                db.session.commit()
            return jsonify({"success": True, "message": f"Your password is: {pwd}"})
    elif user_type == "faculty":
        fac = Faculty.query.filter_by(email=email).first()
        if fac:
            pwd = fac.plain_password
            if not pwd:
                pwd = "fac_" + gen_code(4)
                fac.set_password(pwd)
                db.session.commit()
            return jsonify({"success": True, "message": f"Your password is: {pwd}"})

    return jsonify({"success": False, "message": "Email not found."})

@app.route("/faculty/view_student_pwd", methods=["POST"])
def faculty_view_student_pwd():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    roll_no = data.get("roll_no", "").strip()
    fac_pwd = data.get("faculty_password", "").strip()
    
    fac = Faculty.query.filter_by(faculty_id=session["faculty_id"]).first()
    if not fac or not fac.check_password(fac_pwd):
        return jsonify({"success": False, "message": "Invalid faculty password."})
        
    student = User.query.filter_by(roll_no=roll_no).first()
    if not student:
        return jsonify({"success": False, "message": "Student not found."})
        
    pwd = student.plain_password
    if not pwd:
        pwd = "spas_" + gen_code(4)
        student.set_password(pwd)
        db.session.commit()
        
    return jsonify({"success": True, "password": pwd})


@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if not is_student():
        return redirect(url_for("index"))

    user = User.query.filter_by(roll_no=session["roll_no"]).first()
    
    # Safely match year formats (e.g. "1st" vs "1st Year")
    uy = (user.year or "").strip().lower().replace(" year", "")
    all_subj = Subject.query.all()
    filtered_subj = []
    for s in all_subj:
        sy = (s.year or "").strip().lower().replace(" year", "")
        # Allow exact match, or empty/neutral year fallback for older records
        if sy == uy or not sy or sy == "all":
            filtered_subj.append(s)
            
    subjects: list[dict] = [s.to_dict() for s in filtered_subj]
    
    # If no subjects specifically for this year yet, we can either show empty or all (we show empty to be strict to user rules)
    # But if there are generic subjects with no year, maybe include them too?
    # Requirement: "only display 1st year subject for the 1st year..."
    
    # Get selected subject
    first_subject = subjects[0]['name'] if subjects else 'all'
    session_subject = session.get('selected_subject', first_subject)
    selected_subject = request.args.get('subject', session_subject)
    valid_names = [s['name'] for s in subjects]
    if selected_subject not in valid_names:
        selected_subject = first_subject
    session['selected_subject'] = selected_subject

    # Get practicals
    if selected_subject == 'all':
        practicals = get_all_practicals_flat()
    else:
        practicals = get_all_practicals_for_subject(selected_subject)

    student_results = get_student_results(session["roll_no"])
    submitted_practicals = {r.get("Practical", "") for r in student_results}

    if request.method == "POST":
        practical_name = request.form.get("practical_name")

        if practical_name in submitted_practicals:
            flash("You have already submitted this practical.", "error")
            return redirect(url_for("dashboard", subject=selected_subject))

        questions = [q.to_dict() for q in Question.query.filter_by(practical_name=practical_name).all()]

        if not questions:
            flash("No questions available for this practical. Please contact faculty.", "error")
            return redirect(url_for("dashboard", subject=selected_subject))

        random.shuffle(questions)
        selected = questions[:20]

        for key in ["last_result", "last_result_id", "exam_question_ids",
                     "exam_questions", "exam_start_time", "exam_duration", "practical_name"]:
            session.pop(key, None)

        session["exam_question_ids"] = [int(q["id"]) for q in selected]
        session["exam_start_time"] = datetime.now().timestamp()
        session["exam_duration"] = EXAM_DURATION_SECONDS
        session["practical_name"] = practical_name
        session.modified = True
        return redirect(url_for("exam"))

    # Attendance warning
    att_warning = None
    _roll = session.get("roll_no")
    _subj_counts = {}
    all_sessions = AttendanceSession.query.all()
    for _sess in all_sessions:
        _subj = _sess.subject
        if _subj not in _subj_counts:
            _subj_counts[_subj] = {"present": 0, "total": 0}
        _subj_counts[_subj]["total"] += 1
        rec = AttendanceRecord.query.filter_by(session_id=_sess.id, roll_no=_roll).first()
        if rec:
            _subj_counts[_subj]["present"] += 1
    for _subj, _c in _subj_counts.items():
        if _c["total"] == 0:
            continue
        _pct = _c["present"] / _c["total"] * 100
        if _pct < 75:
            att_warning = {"level": "red", "message": f"Low Attendance: {_subj} attendance {_pct:.0f}% — below 75%."}
            break

    return render_template("dashboard.html", attendance_warning=att_warning, title="Dashboard",
                           user=user.to_dict() if user else {}, practicals=practicals,
                           submitted_practicals=submitted_practicals, student_results=student_results,
                           subjects=subjects, selected_subject=selected_subject)


@app.route("/faculty_dashboard")
def faculty_dashboard():
    if not is_faculty():
        return redirect(url_for("index"))

    fac = Faculty.query.filter_by(faculty_id=session["faculty_id"]).first()
    all_students_q = User.query.all()
    all_students = {u.roll_no: u.to_dict() for u in all_students_q}
    
    all_subjects = [s.to_dict() for s in Subject.query.all()]
    fac_subjects = [s.strip() for s in (fac.subjects or "").split(",") if s.strip()]
    if fac_subjects:
        subjects = [s for s in all_subjects if s["name"] in fac_subjects]
    else:
        subjects = all_subjects

    selected_batch = request.args.get('batch', 'all')
    first_subject = subjects[0]['name'] if subjects else 'all'
    selected_subject = request.args.get('subject', first_subject)

    # Determine the year of the selected subject so we can filter students accordingly
    selected_subject_year = None
    if selected_subject and selected_subject != 'all':
        subj_obj = Subject.query.filter_by(name=selected_subject).first()
        if subj_obj and subj_obj.year:
            # Normalize: strip trailing " Year" so "1st Year" == "1st"
            selected_subject_year = subj_obj.year.strip().lower().replace(" year", "")

    # Filter students by year (if a subject with a known year is selected) and then by batch
    def _student_matches(student_data):
        if selected_subject_year:
            sy = (student_data.get('year') or '').strip().lower().replace(' year', '')
            if sy != selected_subject_year:
                return False
        if selected_batch != 'all':
            if student_data.get('batch', '1') != selected_batch:
                return False
        return True

    students = {k: v for k, v in all_students.items() if _student_matches(v)}

    # Get practicals
    if selected_subject != 'all':
        practicals = get_all_practicals_for_subject(selected_subject)
    else:
        practicals = get_all_practicals_flat()

    all_batches = sorted(set(s.get('batch', '1') for s in all_students.values()))

    # Get all results — latest per (roll_no, practical)
    all_results_q = Result.query.order_by(Result.timestamp.desc()).all()
    _seen = set()
    all_results = []
    for r in all_results_q:
        key = (r.roll_no, r.practical_name)
        if key not in _seen:
            _seen.add(key)
            all_results.append(r.to_dict())

    # Student performance
    student_performance = {}
    for student_id, student_data in students.items():
        student_results = [r for r in all_results if r.get("Roll No") == student_id]
        practical_scores = {}
        total_score: int = 0
        count: int = 0

        for result in student_results:
            practical_name = result.get("Practical", "").strip()
            if not practical_name:
                continue
            practicals_stripped = [p.strip() for p in practicals]
            if selected_subject != 'all' and practical_name not in practicals_stripped:
                continue
            try:
                canonical = practicals[[p.strip() for p in practicals].index(practical_name)]
            except ValueError:
                canonical = practical_name

            score_str = result.get("Score", "0 / 0")
            if '/' in score_str:
                try:
                    score = int(score_str.split('/')[0].strip())
                except ValueError:
                    score = 0
                if canonical not in practical_scores:
                    practical_scores[canonical] = score
                    total_score += score
                    count += 1

        avg_score = round(float(total_score / count), 2) if count > 0 else 0
        student_performance[student_id] = {
            "name": student_data["full_name"],
            "branch": student_data["branch"],
            "year": student_data["year"],
            "batch": student_data.get("batch", "1"),
            "email": student_data["email"],
            "practical_scores": practical_scores,
            "total": total_score,
            "average": avg_score,
            "exams_taken": count
        }

    # Practical submissions
    practical_submissions = {}
    for practical_name in practicals:
        submitted_students = []
        pname_stripped = practical_name.strip()
        for result in all_results:
            if result.get("Practical", "").strip() == pname_stripped:
                roll_no = result.get("Roll No")
                if roll_no and roll_no in students and roll_no not in [s['roll_no'] for s in submitted_students]:
                    student = students.get(roll_no)
                    if student:
                        submitted_students.append({
                            'roll_no': roll_no,
                            'name': student['full_name'],
                            'batch': student.get('batch', '1')
                        })
        practical_submissions[practical_name] = submitted_students

    return render_template("faculty_dashboard.html", title="Faculty Dashboard",
                           faculty=fac.to_dict() if fac else {},
                           students=students, all_students=all_students, practicals=practicals,
                           results=all_results, student_performance=student_performance,
                           practical_submissions=practical_submissions, all_batches=all_batches,
                           subjects=subjects, selected_subject=selected_subject,
                           selected_batch=selected_batch)


@app.route("/api/add_subject", methods=["POST"])
def add_subject():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json()
        subject_name = data.get("name", "").strip()
        year = data.get("year", "1st").strip()
        if not subject_name:
            return jsonify({"success": False, "message": "Subject name required"}), 400
        if Subject.query.filter(db.func.lower(Subject.name) == subject_name.lower()).first():
            return jsonify({"success": False, "message": "Subject already exists"}), 400

        s = Subject(name=subject_name, year=year)
        db.session.add(s)
        db.session.commit()
        return jsonify({"success": True, "subject": {"id": str(s.id), "name": s.name, "year": s.year}}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/add_practical", methods=["POST"])
def add_practical():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json()
        practical_name = data.get("name", "").strip()
        subject_id = data.get("subject_id", "1")
        if not practical_name:
            return jsonify({"success": False, "message": "Practical name required"}), 400

        # Check if exists in any subject
        existing = Practical.query.filter_by(name=practical_name).first()
        if existing:
            return jsonify({"success": False, "message": "Practical already exists"}), 400

        sort_num = extract_practical_number(practical_name)
        subj = Subject.query.get(int(subject_id))
        if not subj:
            return jsonify({"success": False, "message": "Subject not found"}), 404

        p = Practical(name=practical_name, subject_id=subj.id, sort_order=sort_num)
        db.session.add(p)
        db.session.commit()
        return jsonify({"success": True, "practical": practical_name}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove_practical", methods=["POST"])
def remove_practical():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json()
        practical_name = data.get("name", "").strip()
        p = Practical.query.filter_by(name=practical_name).first()
        if not p:
            return jsonify({"success": False, "message": "Practical not found"}), 404

        # Also delete questions linked to this practical
        Question.query.filter_by(practical_id=p.id).delete()
        db.session.delete(p)
        db.session.commit()
        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/export_excel")
def export_excel():
    """Export all student performance to Excel"""
    if not is_faculty():
        return redirect(url_for("index"))
    try:
        all_students = {u.roll_no: u.to_dict() for u in User.query.all()}
        selected_subject = request.args.get('subject', 'all')

        if selected_subject != 'all':
            practicals = get_all_practicals_for_subject(selected_subject)
        else:
            practicals = get_all_practicals_flat()

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Performance"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["Roll No", "Name", "Branch", "Year", "Batch"]
        headers.extend(practicals)
        headers.extend(["Total", "Average"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row = 2
        for roll_no, student in all_students.items():
            ws.cell(row, 1, roll_no).border = border
            ws.cell(row, 2, student["full_name"]).border = border
            ws.cell(row, 3, student["branch"]).border = border
            ws.cell(row, 4, student["year"]).border = border
            ws.cell(row, 5, student.get("batch", "1")).border = border

            student_results = get_student_results(roll_no)
            scores_dict = {}
            for result in student_results:
                practical_name = result.get("Practical")
                if practical_name and practical_name in practicals:
                    score_str = result.get("Score", "0 / 20")
                    if '/' in score_str:
                        scores_dict[practical_name] = int(score_str.split('/')[0].strip())

            col_idx = 6
            total: int = 0
            count: int = 0
            for practical in practicals:
                score = scores_dict.get(practical, None)
                cell = ws.cell(row, col_idx, score if score is not None else "-")
                cell.border = border
                if score is not None:
                    total += score
                    count += 1
                    cell.alignment = Alignment(horizontal="center")
                col_idx += 1

            ws.cell(row, col_idx, total).border = border
            ws.cell(row, col_idx, total).alignment = Alignment(horizontal="center")
            ws.cell(row, col_idx, total).font = Font(bold=True)

            avg = round(float(total / count), 2) if count > 0 else 0
            ws.cell(row, col_idx + 1, avg).border = border
            ws.cell(row, col_idx + 1, avg).alignment = Alignment(horizontal="center")
            ws.cell(row, col_idx + 1, avg).font = Font(bold=True)
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        subject_suffix = f"_{selected_subject}" if selected_subject != 'all' else ""
        filename = f"student_performance{subject_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f"Error exporting to Excel: {str(e)}", "error")
        return redirect(url_for("faculty_dashboard"))


@app.route("/delete_account", methods=["POST"])
def delete_account():
    if is_student():
        roll_no = session["roll_no"]
        user = User.query.filter_by(roll_no=roll_no).first()
        if user:
            Result.query.filter_by(roll_no=roll_no).delete()
            db.session.delete(user)
            db.session.commit()
        session.clear()
        flash("Your account has been deleted successfully.", "success")
        return redirect(url_for("index"))

    elif is_faculty():
        faculty_id = session["faculty_id"]
        fac = Faculty.query.filter_by(faculty_id=faculty_id).first()
        if fac:
            db.session.delete(fac)
            db.session.commit()
        session.clear()
        flash("Your account has been deleted successfully.", "success")
        return redirect(url_for("index"))

    return redirect(url_for("index"))


@app.route("/update_profile", methods=["POST"])
def update_profile():
    if not is_student():
        return redirect(url_for("index"))

    user = User.query.filter_by(roll_no=session["roll_no"]).first()
    if user:
        user.full_name = request.form.get("full_name", "").strip()
        user.branch = request.form.get("branch")
        user.year = request.form.get("year")
        user.batch = request.form.get("batch", "1")
        user.email = request.form.get("email", "").strip()
        db.session.commit()

        session["full_name"] = user.full_name
        session["branch"] = user.branch
        session["year"] = user.year
        session["batch"] = user.batch
        session["email"] = user.email
        flash("Profile updated successfully!", "success")

    return redirect(url_for("dashboard"))


@app.route("/faculty/update_profile", methods=["POST"])
def faculty_update_profile():
    if not is_faculty():
        return redirect(url_for("index"))

    fac = Faculty.query.filter_by(faculty_id=session["faculty_id"]).first()
    if fac:
        fac.full_name = request.form.get("full_name", "").strip()
        fac.department = request.form.get("department")
        fac.email = request.form.get("email", "").strip()
        db.session.commit()

        session["full_name"] = fac.full_name
        session["department"] = fac.department
        session["email"] = fac.email
        flash("Profile updated successfully!", "success")

    return redirect(url_for("faculty_dashboard"))


@app.route("/student/change_password", methods=["POST"])
def student_change_password():
    if not is_student():
        return redirect(url_for("index"))
    
    identifier = request.form.get("identifier", "").strip()
    new_password = request.form.get("new_password", "").strip()
    
    user = User.query.filter_by(roll_no=session["roll_no"]).first()
    if user and (user.roll_no == identifier or user.email == identifier):
        if len(new_password) < 6:
            flash("Password must be at least 6 characters.", "error")
        else:
            user.set_password(new_password)
            db.session.commit()
            flash("Password updated successfully!", "success")
    else:
        flash("Invalid Roll No or Email provided.", "error")
        
    return redirect(url_for("dashboard"))


@app.route("/faculty/change_password", methods=["POST"])
def faculty_change_password():
    if not is_faculty():
        return redirect(url_for("index"))
        
    fac_id = request.form.get("faculty_id", "").strip()
    email = request.form.get("email", "").strip()
    new_password = request.form.get("new_password", "").strip()
    
    fac = Faculty.query.filter_by(faculty_id=session["faculty_id"]).first()
    if fac and fac.faculty_id == fac_id and fac.email == email:
        if len(new_password) < 6:
            flash("Password must be at least 6 characters.", "error")
        else:
            fac.set_password(new_password)
            db.session.commit()
            flash("Password updated successfully!", "success")
    else:
        flash("Invalid Faculty ID or Email provided.", "error")
        
    return redirect(url_for("faculty_dashboard"))


@app.route("/exam")
def exam():
    if not is_student():
        return redirect(url_for("index"))

    if "exam_question_ids" not in session:
        flash("Start exam from dashboard first.", "error")
        return redirect(url_for("dashboard"))

    q_ids = session["exam_question_ids"]
    questions_all = {q.id: q.to_dict() for q in Question.query.filter(Question.id.in_(q_ids)).all()}
    questions = [questions_all[qid] for qid in q_ids if qid in questions_all]

    if not questions:
        flash("Exam data missing. Please start again from the dashboard.", "error")
        return redirect(url_for("dashboard"))

    start_time = session.get("exam_start_time")
    duration = session.get("exam_duration", EXAM_DURATION_SECONDS)
    now_ts = datetime.now().timestamp()
    remaining = int(start_time + duration - now_ts)

    if remaining <= 0:
        return redirect(url_for("submit_exam"))

    practical_name = session.get("practical_name", "")
    return render_template("exam.html", questions=questions, remaining=remaining, practical_name=practical_name)


@app.route("/submit_exam", methods=["POST", "GET"])
def submit_exam():
    if not is_student():
        return redirect(url_for("index"))

    if "exam_question_ids" not in session:
        return redirect(url_for("dashboard"))

    q_ids = session["exam_question_ids"]
    questions_all = {q.id: q.to_dict() for q in Question.query.filter(Question.id.in_(q_ids)).all()}
    questions = [questions_all[qid] for qid in q_ids if qid in questions_all]

    if not questions:
        return redirect(url_for("dashboard"))

    submitted_answers = {}
    if request.method == "POST":
        for q in questions:
            qid = str(q["id"])
            submitted_answers[qid] = request.form.get(f"answer_{qid}")

    practical_name = (request.form.get("practical_name") or "").strip() or session.get("practical_name", "")

    total = len(questions)
    attempted: int = 0
    correct: int = 0
    detailed_answers = []

    for q in questions:
        qid = str(q["id"])
        ans = submitted_answers.get(qid)
        is_correct = ans == q["answer"] if ans else False

        if ans:
            attempted += 1
            if is_correct:
                correct += 1

        detailed_answers.append({
            "question": q["question"],
            "options": q.get("options", {}),
            "student_answer": ans if ans else "NOT ATTEMPTED",
            "correct_answer": q["answer"],
            "status": "CORRECT" if is_correct else ("WRONG" if ans else "NOT ATTEMPTED")
        })

    wrong = total - correct
    score = correct

    user = User.query.filter_by(roll_no=session["roll_no"]).first()
    dt_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts = int(datetime.now().timestamp())

    # Save result to database
    r = Result(
        roll_no=user.roll_no,
        name=user.full_name,
        branch=user.branch,
        year=user.year,
        batch=user.batch,
        email=user.email,
        practical_name=practical_name,
        score=f"{score} / {total}",
        attempted=attempted,
        correct=correct,
        wrong=wrong,
        datetime_str=dt_str,
        timestamp=ts,
        details_json=json.dumps(detailed_answers, ensure_ascii=False),
    )
    db.session.add(r)
    db.session.commit()

    session["last_result"] = {
        "roll_no": user.roll_no,
        "name": user.full_name,
        "branch": user.branch,
        "year": user.year,
        "batch": user.batch,
        "email": user.email,
        "practical_name": practical_name,
        "score": f"{score} / {total}",
        "total_questions": total,
        "attempted": attempted,
        "correct": correct,
        "wrong": wrong,
        "datetime": dt_str,
        "detailed_answers": [],
    }
    session["last_result_id"] = r.id

    session.pop("exam_question_ids", None)
    session.pop("exam_start_time", None)
    session.pop("exam_duration", None)
    session.pop("practical_name", None)

    return redirect(url_for("result"))


@app.route("/result")
def result():
    result_data = session.get("last_result")
    result_id = session.get("last_result_id")

    if not result_data:
        flash("No result found. Please take an exam first.", "error")
        return redirect(url_for("dashboard"))

    if not result_data.get("detailed_answers") and result_id:
        r = Result.query.get(result_id)
        if r:
            result_data["detailed_answers"] = json.loads(r.details_json) if r.details_json else []

    return render_template("result.html", result=result_data, filename=None)


@app.route("/view_result/<practical_name>")
def view_result(practical_name):
    if not is_student():
        return redirect(url_for("index"))

    r = Result.query.filter_by(roll_no=session['roll_no'], practical_name=practical_name)\
        .order_by(Result.timestamp.desc()).first()

    if not r:
        flash("Result not found.", "error")
        return redirect(url_for("dashboard"))

    result_data = r.to_full_dict()
    return render_template("result.html", result=result_data, filename=None)


@app.route("/download/<int:result_id>")
def download(result_id):
    """Download result as txt file"""
    r = Result.query.get(result_id)
    if not r:
        flash("Result not found.", "error")
        return redirect(url_for("dashboard"))

    txt_content = r.to_txt()
    filename = f"Result_RollNo_{r.roll_no}_{r.timestamp}.txt"
    return Response(
        txt_content,
        mimetype='text/plain',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("index"))


@app.route("/faculty/view_result/<roll_no>/<practical_name>")
def faculty_view_result(roll_no, practical_name):
    if not is_faculty():
        return redirect(url_for("index"))

    r = Result.query.filter_by(roll_no=roll_no, practical_name=practical_name)\
        .order_by(Result.timestamp.desc()).first()

    result_data = r.to_full_dict() if r else None
    return render_template("result.html", result=result_data, filename=None, is_faculty_view=True)


@app.route("/faculty/get_result_data/<roll_no>/<practical_name>")
def get_result_data(roll_no, practical_name):
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    r = Result.query.filter_by(roll_no=roll_no, practical_name=practical_name)\
        .order_by(Result.timestamp.desc()).first()

    if r:
        return jsonify({"success": True, "result": r.to_full_dict()})
    return jsonify({"success": False, "message": "Result not found"})


@app.route("/faculty/get_result_txt/<roll_no>/<practical_name>")
def get_result_txt(roll_no, practical_name):
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    r = Result.query.filter_by(roll_no=roll_no, practical_name=practical_name)\
        .order_by(Result.timestamp.desc()).first()

    if r:
        return jsonify({
            "success": True,
            "content": r.to_txt(),
            "filename": f"Result_RollNo_{r.roll_no}_{r.timestamp}.txt"
        })
    return jsonify({"success": False, "message": "File not found"})


@app.route("/api/get_questions")
def get_questions():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    practical_name = request.args.get("practical", "").strip()
    if not practical_name:
        return jsonify({"success": False, "message": "Practical name required"}), 400
    questions = [q.to_dict() for q in Question.query.filter_by(practical_name=practical_name).all()]
    return jsonify({"success": True, "questions": questions, "count": len(questions)})


@app.route("/api/add_question", methods=["POST"])
def add_question():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json()
        practical_name = data.get("practical", "").strip()
        question_text = data.get("question", "").strip()
        options = data.get("options", {})
        answer = data.get("answer", "").strip()

        if not practical_name:
            return jsonify({"success": False, "message": "Practical name required"}), 400
        if not question_text:
            return jsonify({"success": False, "message": "Question text required"}), 400
        if not all(options.get(k, "").strip() for k in ["A", "B", "C", "D"]):
            return jsonify({"success": False, "message": "All 4 options are required"}), 400
        if answer not in ["A", "B", "C", "D"]:
            return jsonify({"success": False, "message": "Correct answer must be A, B, C or D"}), 400

        existing_count = Question.query.filter_by(practical_name=practical_name).count()
        if existing_count >= 20:
            return jsonify({"success": False, "message": "Maximum 20 questions allowed per practical"}), 400

        prac = Practical.query.filter_by(name=practical_name).first()
        if not prac:
            return jsonify({"success": False, "message": "Practical not found"}), 404

        q = Question(
            practical_id=prac.id,
            practical_name=practical_name,
            question=question_text,
            option_a=options.get("A", "").strip(),
            option_b=options.get("B", "").strip(),
            option_c=options.get("C", "").strip(),
            option_d=options.get("D", "").strip(),
            answer=answer,
        )
        db.session.add(q)
        db.session.commit()
        return jsonify({"success": True, "question": q.to_dict(), "total": existing_count + 1}), 200
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/delete_question", methods=["POST"])
def delete_question():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json()
        question_id = data.get("id")
        q = Question.query.get(question_id)
        if not q:
            return jsonify({"success": False, "message": "Question not found"}), 404
        db.session.delete(q)
        db.session.commit()
        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


# ═══════════════════════════════════════════════════════════
# ATTENDANCE SYSTEM
# ═══════════════════════════════════════════════════════════

@app.route("/api/generate_attendance_code", methods=["POST"])
def generate_attendance_code():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    data = request.get_json()
    subject = data.get("subject", "").strip()
    batch = data.get("batch", "all").strip()
    seconds = int(data.get("valid_seconds", 600))
    if not subject:
        return jsonify({"success": False, "message": "Subject required"})

    code = gen_code()
    now = int(_time.time())
    today = _date.today().isoformat()
    sid = f"sess_{now}_{subject}_{batch}"

    a = AttendanceSession(
        session_key=sid, code=code, subject=subject, batch=batch,
        date=today, created_at=now, valid_seconds=seconds
    )
    db.session.add(a)
    db.session.commit()

    return jsonify({"success": True, "code": code, "session_id": sid,
                    "valid_seconds": seconds, "date": today})


@app.route("/api/mark_attendance", methods=["POST"])
def mark_attendance():
    if not is_student():
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    roll_no = session.get("roll_no")
    data = request.get_json()
    code = data.get("code", "").strip().upper()
    if not code:
        return jsonify({"success": False, "message": "Please enter a code"})

    now = int(_time.time())
    att_sess = AttendanceSession.query.filter_by(code=code).first()
    if not att_sess:
        return jsonify({"success": False, "message": "Invalid code. Please try again."})

    elapsed_seconds = now - att_sess.created_at
    if elapsed_seconds > att_sess.valid_seconds:
        return jsonify({"success": False, "message": "Code expired. Ask faculty for a new one."})

    existing = AttendanceRecord.query.filter_by(session_id=att_sess.id, roll_no=roll_no).first()
    if existing:
        return jsonify({"success": False, "message": "Attendance already marked for this session."})

    user = User.query.filter_by(roll_no=roll_no).first()
    rec = AttendanceRecord(
        session_id=att_sess.id, roll_no=roll_no,
        name=user.full_name if user else roll_no,
        marked_at=now
    )
    db.session.add(rec)
    db.session.commit()

    return jsonify({"success": True,
                    "message": f'Attendance marked for {att_sess.subject} on {att_sess.date}'})


@app.route("/api/get_attendance")
def get_attendance_api():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    subject = request.args.get("subject", "")
    batch = request.args.get("batch", "all")
    date = request.args.get("date", "")

    query = AttendanceSession.query
    if subject:
        query = query.filter_by(subject=subject)
    if batch and batch != "all":
        query = query.filter_by(batch=batch)
    if date:
        query = query.filter_by(date=date)

    sessions = []
    for sess in query.order_by(AttendanceSession.created_at.desc()).all():
        records = AttendanceRecord.query.filter_by(session_id=sess.id).all()
        sessions.append({
            "session_id": sess.session_key,
            "subject": sess.subject,
            "batch": sess.batch,
            "date": sess.date,
            "valid_seconds": sess.valid_seconds,
            "created_at": sess.created_at,
            "total_marked": len(records),
            "marked_students": [{"name": r.name, "roll_no": r.roll_no, "marked_at": r.marked_at} for r in records]
        })

    return jsonify({"success": True, "sessions": sessions})


@app.route("/api/delete_attendance_session", methods=["POST"])
def delete_attendance_session():
    """Faculty: delete an attendance session and all its records."""
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    try:
        data = request.get_json()
        session_key = (data.get("session_id") or "").strip()
        if not session_key:
            return jsonify({"success": False, "message": "session_id required"}), 400

        att_sess = AttendanceSession.query.filter_by(session_key=session_key).first()
        if not att_sess:
            return jsonify({"success": False, "message": "Session not found"}), 404

        # Delete records first, then the session
        AttendanceRecord.query.filter_by(session_id=att_sess.id).delete()
        db.session.delete(att_sess)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/get_my_attendance")
def get_my_attendance():
    if not is_student():
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    roll_no = session.get("roll_no")
    all_sessions = AttendanceSession.query.order_by(AttendanceSession.created_at.desc()).all()
    counts = {}
    session_details = []

    for sess in all_sessions:
        subj = sess.subject
        if subj not in counts:
            counts[subj] = {"present": 0, "total": 0}
        counts[subj]["total"] += 1
        rec = AttendanceRecord.query.filter_by(session_id=sess.id, roll_no=roll_no).first()
        is_present = rec is not None
        if is_present:
            counts[subj]["present"] += 1
        session_details.append({
            "session_key": sess.session_key,
            "subject": sess.subject,
            "batch": sess.batch,
            "date": sess.date,
            "status": "Present" if is_present else "Absent"
        })

    result_list = []
    for subj, c in counts.items():
        pct = round(c["present"] / c["total"] * 100, 1) if c["total"] else 0
        result_list.append({"subject": subj, "present": c["present"],
                            "total": c["total"], "pct": pct})

    return jsonify({"success": True, "subjects": result_list, "sessions": session_details})


@app.route("/api/download_my_session_attendance")
def download_my_session_attendance():
    """Download a student's attendance record for a specific session as Excel."""
    if not is_student():
        return redirect(url_for("index"))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    session_key = request.args.get("session_id", "").strip()
    roll_no = session.get("roll_no")

    if not session_key:
        flash("Session ID required.", "error")
        return redirect(url_for("dashboard"))

    att_sess = AttendanceSession.query.filter_by(session_key=session_key).first()
    if not att_sess:
        flash("Session not found.", "error")
        return redirect(url_for("dashboard"))

    rec = AttendanceRecord.query.filter_by(session_id=att_sess.id, roll_no=roll_no).first()
    is_present = rec is not None
    user = User.query.filter_by(roll_no=roll_no).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Attendance"

    hf   = Font(bold=True, color="FFFFFF", size=11)
    hfil = PatternFill("solid", fgColor="1565C0")
    thin = Side(border_style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")

    headers = ["Roll No", "Name", "Date", "Subject", "Batch", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfil; c.alignment = ctr; c.border = bdr

    status_val = "Present" if is_present else "Absent"
    row_data = [
        roll_no,
        user.full_name if user else roll_no,
        att_sess.date,
        att_sess.subject,
        att_sess.batch,
        status_val
    ]
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=ci, value=val)
        cell.border = bdr
        cell.alignment = ctr
        if ci == 6:
            cell.fill = PatternFill("solid", fgColor="C8E6C9" if is_present else "FFCDD2")
            cell.font = Font(bold=True, color="1B5E20" if is_present else "B71C1C")

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 35)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe_subj = att_sess.subject.replace(" ", "_")
    fname = f"Attendance_{roll_no}_{safe_subj}_{att_sess.date}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export_attendance_excel")
def export_attendance_excel():
    if not is_faculty():
        return redirect(url_for("index"))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    subject = request.args.get("subject", "")
    batch = request.args.get("batch", "all")
    date = request.args.get("date", "")

    query = AttendanceSession.query
    if subject:
        query = query.filter_by(subject=subject)
    if batch and batch != "all":
        query = query.filter_by(batch=batch)
    if date:
        query = query.filter_by(date=date)

    sessions = [(s.id, s) for s in query.order_by(AttendanceSession.created_at).all()]

    all_users = {u.roll_no: u.to_dict() for u in User.query.all()}
    if batch != "all":
        students = sorted(r for r, u in all_users.items() if u.get("batch") == batch)
    else:
        students = sorted(all_users.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfil = PatternFill("solid", fgColor="1565C0")
    thin = Side(border_style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")

    hdrs = ["Roll No", "Name", "Branch", "Batch"]
    for _, s in sessions:
        hdrs.append(f'{s.date} | {s.subject} | Batch {s.batch}')
    hdrs += ["Present", "Absent", "Attendance %"]

    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfil; c.alignment = ctr; c.border = bdr

    for ri, roll in enumerate(students, 2):
        u = all_users.get(roll, {})
        ws.cell(row=ri, column=1, value=roll).border = bdr
        ws.cell(row=ri, column=2, value=u.get("full_name", "")).border = bdr
        ws.cell(row=ri, column=3, value=u.get("branch", "")).border = bdr
        ws.cell(row=ri, column=4, value=u.get("batch", "")).border = bdr
        present: int = 0
        for ci, (sess_id, s) in enumerate(sessions, 5):
            rec = AttendanceRecord.query.filter_by(session_id=sess_id, roll_no=roll).first()
            val = "P" if rec else "A"
            c2 = ws.cell(row=ri, column=ci, value=val)
            c2.alignment = ctr; c2.border = bdr
            c2.fill = PatternFill("solid", fgColor="C8E6C9" if val == "P" else "FFCDD2")
            if val == "P":
                present += 1
        tot = len(sessions)
        pct = round(present / tot * 100, 1) if tot else 0
        ws.cell(row=ri, column=5 + tot, value=present).border = bdr
        ws.cell(row=ri, column=6 + tot, value=tot - present).border = bdr
        pc = ws.cell(row=ri, column=7 + tot, value=f"{pct}%")
        pc.border = bdr; pc.alignment = ctr
        pc.fill = PatternFill("solid", fgColor="A5D6A7" if pct >= 75 else ("FFE082" if pct >= 60 else "EF9A9A"))

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 35)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'Attendance_{subject or "All"}_{date or "All"}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)